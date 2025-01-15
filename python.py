import openpyxl  # Import library to work with Excel files
import requests  # Import library to make HTTP requests
from bs4 import BeautifulSoup  # Import library to parse HTML content
import time  # Import library to add delays in the script
import urllib.parse  # Import library to handle URL encoding

# Function to fetch Google search results for a given query
def fetch_google_results(query, start_index=0):
    # Create the search URL using the query and starting index for pagination
    search_url = "https://www.google.com/search?q=" + urllib.parse.quote(query) + "&start=" + str(start_index)
    
    # Set up the headers to mimic a web browser (helps avoid being blocked by Google)
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    }
    
    # Make the HTTP request to fetch the search results page
    response = requests.get(search_url, headers=headers)
    
    # If the request was not successful, print an error message and return an empty list
    if response.status_code != 200:
        print(f"Failed to fetch search results for query '{query}'")
        return []
    
    # Debug: Print the first 500 characters of the HTML content (for inspection)
    print(f"Raw HTML for query '{query}':\n{response.text[:500]}...")  # Only print the first 500 characters

    # Parse the response content using BeautifulSoup to extract useful data
    soup = BeautifulSoup(response.text, "html.parser")
    links = []  # List to store the URLs of the search results
    
    # Look for all anchor tags (<a>) that have an href attribute (which holds the URLs)
    for a_tag in soup.find_all('a', href=True):
        href = a_tag['href']
        
        # If the href contains "url?q=", it is a valid search result URL
        if "url?q=" in href:
            # Extract the actual URL from the href attribute
            link = href.split("url?q=")[1].split("&")[0]
            links.append(link)  # Add the extracted URL to the links list
    
    # Debug: Print the number of URLs found and the URLs themselves
    print(f"Found {len(links)} URLs for query '{query}':")
    for url in links:
        print(f"  {url}")
    
    return links  # Return the list of URLs

# Function to search for keywords in an Excel file and update it with the search result URLs
def search_and_update_excel(file_name):
    try:
        # Load the existing Excel workbook
        workbook = openpyxl.load_workbook(file_name)
        
        # Select the "Google Search Results" sheet in the workbook
        sheet = workbook["Google Search Results"]

        # Loop through each row in the sheet, starting from row 2 (skip header in row 1)
        for row_index in range(2, sheet.max_row + 1):
            keyword_cell = sheet.cell(row=row_index, column=1)  # Get the keyword from column 1
            keyword = keyword_cell.value  # Store the keyword as a string

            # Strip any unnecessary spaces or characters from the keyword
            if keyword:
                keyword = keyword.strip()

            # If the keyword is empty, skip the row and move to the next one
            if not keyword:
                print(f"Row {row_index} is empty. Skipping...")
                continue
            
            print(f"Searching for: {keyword}")
            
            try:
                urls = []  # List to store the found URLs for the current keyword
                # Loop to fetch search results from multiple pages (up to 4 pages of 10 results each)
                for start_index in range(0, 40, 10):  # Fetch results for 4 pages (start_index 0, 10, 20, 30)
                    result_urls = fetch_google_results(keyword, start_index)  # Get the URLs for the current page
                    urls.extend(result_urls)  # Add the results to the urls list
                    
                    # Wait for 5 seconds before fetching the next set of results to avoid being blocked
                    print(f"Waiting for a few seconds before fetching the next page of results...")
                    time.sleep(5)
                
                # Print each found URL in the console (for debugging purposes)
                for url in urls:
                    print(f"Found URL: {url}")
                
                # Write the found URLs into the Excel sheet, starting from column B
                for col_index, url in enumerate(urls, start=2):
                    sheet.cell(row=row_index, column=col_index).value = url
                
                print(f"Results written for keyword: {keyword}")
            except Exception as e:
                print(f"Error fetching results for '{keyword}': {e}")
        
        # Save the updated Excel file with the new search results
        workbook.save(file_name)
        print(f"All updates saved to {file_name}")
    except FileNotFoundError:
        print(f"File {file_name} not found. Please check the file path.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Main execution of the script
if __name__ == "__main__":
    excel_file = "GoogleSearchResults.xlsx"  # Specify the Excel file to update
    search_and_update_excel(excel_file)  # Run the function to search and update the Excel file
